import openpyxl


class HomeData:

    test_home_data = [{'firstname': 'Rahul', 'lastname': 'Shetty', 'gender': 'Male'}]

    @staticmethod
    def get_test_data(test_case):
        d = dict()
        wb = openpyxl.load_workbook('D:\\ReadData.xlsx')
        sheet = wb.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case:
                for j in range(2, sheet.max_column + 1):
                    d[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [d]